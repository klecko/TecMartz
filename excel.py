import openpyxl
import sys


def get_columna_existencias(sheet):
    for i, col in enumerate(sheet.iter_cols()):
        if "EXIST" in str(col[0].value):
            return i
    return -1

def get_modelos(path, numero_disenos):
    
    wb = openpyxl.load_workbook(path, data_only=True) #carga el excel
    #en "modelos" es donde se guardará todo
    modelos = {} #keys: modelo. values: list of dicts (keys: diseño, existencias)
    errores = []

    for sheet in wb.worksheets[:numero_disenos]: #por cada hoja en las NUMERO_DISENOS primeras hojas
        col_existencias = get_columna_existencias(sheet) #obtiene la columna donde se encuentran las existencias
        #si no encuentra la columna, error
        if col_existencias == -1: errores.append("ERROR OBTENIENDO LA COLUMNA DE LAS EXISTENCIAS EN HOJA " + sheet.title); continue

        marca = "" 
        for row in sheet.iter_rows(min_row=2): #por cada fila dentro de la hoja
            if row[1].value == None: continue #si no hay modelo, pasa a la siguiente fila
            if row[0].value == None: #si no hay nada en la primera columna, es marca
                marca = str(row[1].value) + " " #asigna el valor de la segunda columna a la variable "marca"
            else: #si hay algo en la primera columna (precio), es modelo
                modelo = marca + str(row[1].value) #junta la marca y el valor de la segunda columna y lo guarda
                try: #intenta obtener el valor de las existencias
                    existencias = int(row[col_existencias].value)
                except: #si ocurre un error, se imprime un mensaje y se pasa a la siguiente fila
                    errores.append("ERROR OBTENIENDO EXISTENCIAS. DISEÑO: " + str(sheet.title) + ". MODELO: " + str(modelo))
                    continue

                #crea un diccionario con los datos diseño y existencias
                d = {"diseño":sheet.title, "existencias":existencias} 
                if modelos.get(modelo): #si ya existía el modelo actual, le añade el nuevo diseño encontrado
                    modelos[modelo].append(d)
                else: #si no existía, simplemente lo crea
                    modelos[modelo] = [d]
    
    return modelos, errores


def buscar_modelos(modelos, name):
    result = {}
    name = name.lower().replace(" ", "")
    for modelo in modelos.keys():
        if name in modelo.lower().replace(" ", ""):
            result[modelo] = modelos[modelo]
    return result


def sprint_modelos(modelos):
    result = ""
    for k, v in modelos.items():
        result += "Modelo: " + str(k) + "\n"
        for d in v:
            result += "Diseño " + str(d["diseño"]) + " con " + str(d["existencias"]) + " existencias." + "\n"
        result += "\n"
    return result

""" path = "excel/Existencias actual 070719.xlsx"
modelos = get_modelos(path)

search = "Mate 20 lite"
print("BÚSQUEDA:", search, "\n")
d = buscar_modelos(modelos, search)
for k, v in d.items():
    print("Modelo:", k)
    for d2 in v:
        print("Diseño", d2["diseño"], "con", d2["existencias"], "existencias.")
    print()
 """
